import argparse
import os
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import psycopg2
from psycopg2.extras import execute_batch
from openpyxl import load_workbook

from config import settings

USD_TO_DOP = Decimal("66")
ITBIS_RATE = Decimal("0.18")
CONVERSION_FACTOR = USD_TO_DOP * (Decimal("1") + ITBIS_RATE)


def parse_decimal(value):
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None

    text = (
        text.replace("USD", "")
        .replace("UST", "")
        .replace("US$", "")
        .replace("$", "")
        .replace("RD$", "")
        .replace(" ", "")
    )

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def ensure_lista_de_precios_table(conn):
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS lista_de_precios (
                id SERIAL PRIMARY KEY,
                codigo VARCHAR(80) NOT NULL UNIQUE,
                descripcion TEXT NOT NULL,
                precio_usd NUMERIC(14, 4) NOT NULL,
                precio_dop NUMERIC(14, 2) NOT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_lista_de_precios_codigo ON lista_de_precios(codigo)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_lista_de_precios_descripcion ON lista_de_precios(LOWER(descripcion))")


def load_rows_from_excel(excel_path, sheet_name=None):
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = []
    skipped = 0

    for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            skipped += 1
            continue

        codigo = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        descripcion = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        precio_usd = parse_decimal(row[2] if len(row) > 2 else None)

        if index == 1 and codigo.lower() in {"codigo", "code"}:
            continue

        if not codigo and not descripcion and precio_usd is None:
            skipped += 1
            continue

        if not codigo or not descripcion or precio_usd is None:
            skipped += 1
            continue

        precio_dop = (precio_usd * CONVERSION_FACTOR).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        rows.append((codigo, descripcion, precio_usd, precio_dop))

    wb.close()
    return rows, skipped, ws.title


def upsert_rows(conn, rows):
    sql = """
        INSERT INTO lista_de_precios (codigo, descripcion, precio_usd, precio_dop, updated_at)
        VALUES (%s, %s, %s, %s, NOW())
        ON CONFLICT (codigo)
        DO UPDATE SET
            descripcion = EXCLUDED.descripcion,
            precio_usd = EXCLUDED.precio_usd,
            precio_dop = EXCLUDED.precio_dop,
            updated_at = NOW()
    """
    with conn.cursor() as cur:
        execute_batch(cur, sql, rows, page_size=500)


def main():
    parser = argparse.ArgumentParser(description="Importa lista_de_precios desde Excel y convierte USD a DOP con ITBIS.")
    parser.add_argument(
        "--file",
        default=os.path.join(os.path.dirname(__file__), "Lista de Precio POS 2.xlsx"),
        help="Ruta del archivo Excel.",
    )
    parser.add_argument("--sheet", default=None, help="Nombre de la hoja (opcional).")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        raise FileNotFoundError(f"No se encontro el archivo Excel: {args.file}")

    rows, skipped, sheet_used = load_rows_from_excel(args.file, args.sheet)
    if not rows:
        print("No hay filas validas para importar.")
        return

    conn = psycopg2.connect(
        host=settings.DB_HOST,
        port=settings.DB_PORT,
        database=settings.DB_NAME,
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        sslmode="require",
        connect_timeout=int(settings.DB_CONNECT_TIMEOUT),
    )

    try:
        ensure_lista_de_precios_table(conn)
        upsert_rows(conn, rows)
        conn.commit()
    finally:
        conn.close()

    print(f"Hoja usada: {sheet_used}")
    print(f"Filas importadas/actualizadas: {len(rows)}")
    print(f"Filas omitidas: {skipped}")
    print(f"Factor de conversion aplicado (USD->DOP+ITBIS): {CONVERSION_FACTOR}")


if __name__ == "__main__":
    main()
